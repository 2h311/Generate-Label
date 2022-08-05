import string
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

class BaseXlsx:
	def __init__(self, filename: str):
		self.path = Path(filename)

	def check_filename(self) -> None:
		# check if the file exists
		if self.path.exists():
			# check if file is an excel file
			if self.file_type not in self.path.name:
				self.filename += self.file_type
		else:
			raise Exception("File Not Found")

class XlsxReader(BaseXlsx):
	def __init__(self, filename: str, fields: List):
		super().__init__(filename)
		self.file_type = '.xlsx'
		self.fields = fields
		self.letters = string.ascii_uppercase
		#
		self.check_filename()
		self.worksheet = self.open_last_sheet()
		#
		self._max_row = self.worksheet.max_row
		self._min_row = self.worksheet.min_row
		self._max_column = self.worksheet.max_column
		self._min_column = self.worksheet.min_column

	def open_last_sheet(self):
		if (workbook := load_workbook(self.path.name)):
			workbook = workbook.worksheets[-1]
		return workbook

	def grab_headers(self) -> Dict:
		cell_dict = dict()
		for number in range(self._max_column):
			cell = f"{(alphabet := self.letters[number])}{self._min_row}"
			value = self.worksheet[cell].value
			# print(cell, value)
			cell_dict[value] = alphabet
		return cell_dict

	def grab_excel_body(self) -> List:
		header_letters: Dict = self.grab_headers()
		list_ = list()
		# start from 2, skip the first row which serves as headings
		for digit in range(2, self._max_row + 1):
			# check if the value cell of every row has a value
			if self.worksheet[f'A{digit}'].value:
				row_dict = {self.fields[num]: self.worksheet[f"{header_letters[self.fields[num]]}{digit}"].value for num in range(len(self.fields)) }
				print(row_dict)
				list_.append(row_dict)
		return list_

