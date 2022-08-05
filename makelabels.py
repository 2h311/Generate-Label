import io
import string
import tempfile
from pathlib import Path
from typing import Dict, List

import requests
from fake_useragent import UserAgent
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook

class BaseXlsx:
	def __init__(self, filename: str):
		self.path = Path(filename)

	def check_filename(self) -> None:
		# check if the file exists
		if self.path.exists():
			# check if file is an excel file
			if self.file_type not in self.path.name:
				self.filename += self.file_type
		else:
			raise Exception("File Not Found")

class XlsxReader(BaseXlsx):
	def __init__(self, filename: str, fields: List):
		super().__init__(filename)
		self.file_type = '.xlsx'
		self.fields = fields
		self.letters = string.ascii_uppercase
		#
		self.check_filename()
		self.worksheet = self.open_last_sheet()
		#
		self._max_row = self.worksheet.max_row
		self._min_row = self.worksheet.min_row
		self._max_column = self.worksheet.max_column
		self._min_column = self.worksheet.min_column

	def open_last_sheet(self):
		if (workbook := load_workbook(self.path.name)):
			workbook = workbook.worksheets[-1]
		return workbook

	def grab_headers(self) -> Dict:
		cell_dict = dict()
		for number in range(self._max_column):
			cell = f"{(alphabet := self.letters[number])}{self._min_row}"
			value = self.worksheet[cell].value
			# print(cell, value)
			cell_dict[value] = alphabet
		return cell_dict

	def grab_excel_body(self) -> List:
		header_letters: Dict = self.grab_headers()
		list_ = list()
		# start from 2, skip the first row which serves as headings
		for digit in range(2, self._max_row + 1):
			# check if the value cell of every row has a value
			if self.worksheet[f'A{digit}'].value:
				row_dict = {self.fields[num]: self.worksheet[f"{header_letters[self.fields[num]]}{digit}"].value for num in range(len(self.fields)) }
				# print(row_dict)
				list_.append(row_dict)
		return list_

def get_barcode(barcode_param: str):
	url = f"https://barcode.tec-it.com/barcode.ashx?data={barcode_param}&code=EANUCC128&translate-esc=true"
	res = requests.get(url, headers={'user-agent': ua.random}, timeout=45, stream=True)
	if res.ok:
		with tempfile.SpooledTemporaryFile(max_size=1e9) as buffer:
			for chunk in res.iter_content(chunk_size=1024):
			    buffer.write(chunk)
			
			# write image to a pillow object 
			buffer.seek(0)
			image = Image.open(io.BytesIO(buffer.read()))
	return image

def draw_text_on_barcode(barcode_text, barcode_param):
	barcode_image = get_barcode(barcode_param)
	barcode_image = barcode_image.resize((barcode_image.width+350, barcode_image.height+180))

	width, height = barcode_image.size
	margin = 25
	new_height = height + (2*margin)

	# empty image for code and text - it needs margins for text
	new_image = Image.new('RGB', (width+60, new_height), (255, 255, 255))
	new_image.paste(barcode_image, (0, margin))

	# object to draw text
	draw = ImageDraw.Draw(new_image)

	# create cover for the barcode number
	cover_image = Image.new('RGB', (width, 72), color='white')
	new_image.paste(cover_image, (0, 245))	 
	
	# put space in between the barcode text
	display_text = ' '.join([(barcode_text[i:i+4]) for i in range(0, len(barcode_text), 4)])

	# draw text
	fnt = ImageFont.truetype("./fonts/arialbd.ttf", 52)
	draw.text((100, new_height - 70), display_text, fill=(0, 0, 0), font=fnt)   	
	return new_image

def cover_weight(weight: str) -> None: 
	weight_cover_image = Image.new(mode, (60, 70), color=color)
	im.paste(weight_cover_image, (420, 110))

	# set the x coordinate to place the weight according to the length of the weight text
	if len(weight) == 1:
		weight_x_coordinate = 466
	elif len(weight) == 2:
		weight_x_coordinate = 446
	else:
		weight_x_coordinate = 430
	draw.text((weight_x_coordinate, 104), weight, fill=9, font=font)

def cover_date(date: str) -> None:
	date = date.split()[0]
	# cover the date
	date_cover_image = Image.new(mode, (160, 40), color=color)
	im.paste(date_cover_image, (1025, 280))
	draw.text((950, 280), date, fill=9, font=font)

def cover_sender(sender: str) -> None:
	sender = sender.upper()
	sender_list = [string.strip() for string in sender.split(',')]
	sedner = f"{' '.join(sender_list[:-3])}\n{' '.join(sender_list[-3:])}" 

	sender_cover_image = Image.new(mode, (1000, 250), color=color)
	im.paste(sender_cover_image, (70, 460))
	draw.text((70, 460), sedner, fill=9, font=ImageFont.truetype(font_path.as_posix(), 62))

def cover_receiver(receiver: str):
	receiver = receiver.upper()
	# cover the receiver
	receiver_cover_image = Image.new(mode, (950, 240), color=color)
	im.paste(receiver_cover_image, (220, 715))
	draw.text((220, 715), receiver, fill=9, font=ImageFont.truetype(font_path.as_posix(), 62))

def cover_orderid(orderid: str):
	orderid_cover_iamge = Image.new(mode, (400, 72), color=color)
	im.paste(orderid_cover_iamge, (220, 1650))
	draw.text((220, 1676), orderid, fill=9, font=ImageFont.truetype(font_path.as_posix(), 47))

def cover_sku(sku: str):
	# cover the sku
	sku_cover_image = Image.new(mode, (420, 72), color=color)
	im.paste(sku_cover_image, (760, 1650))
	draw.text((800, 1655), sku, fill=9, font=ImageFont.truetype(r'fonts/arialbd.ttf', 60))

def generate_label_from_dict(dict_):
	cover_weight(dict_.get('weight'))
	cover_date(dict_.get('order create time'))
	cover_sender(dict_.get('sender'))
	receiver = dict_.get("First name") + " " + dict_.get("Last name")
	cover_receiver(receiver)
	orderid = dict_.get('Order ID')
	cover_orderid(orderid)
	cover_sku(dict_.get('SKU'))

	ai1 = '420'
	zipp = dict_['Postal code']
	barcode_text = dict_['Tracking number']

	barcode_param = f"({ai1}){zipp}({barcode_text[:2]}){barcode_text[2:]}"
	barcode_image = draw_text_on_barcode(barcode_text, barcode_param)

	# paste barcode 
	im.paste(barcode_image, (100, 1210))
	im.save(f"{barcode_text}.png")

def main():
	excel_filename = "order templat.xlsx"
	fields = [
		'sender',
		'Last name',
		'First name',
		'order create time',
		'weight',
		'Order ID',
		'SKU',
		'Postal code',
		'Tracking number',
	]
	reader = XlsxReader(filename=excel_filename, fields=fields)
	data = reader.grab_excel_body()
	for dict_ in data:
		print(dict_)
		generate_label_from_dict(dict_)

ua = UserAgent(fallback="Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36")
im = Image.open("TMPLATE.png").convert('RGB')
draw = ImageDraw.Draw(im)
mode, color = im.mode, 'white' 
font_path = Path.cwd() / Path('fonts') / Path("arial.ttf")
font = ImageFont.truetype(font_path.as_posix(), 37)
main()